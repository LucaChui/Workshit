# -*- coding: utf-8 -*-
"""
DAM 资产清单导出脚本（默认：抓取全量资产 + 下载图片 + 在 assets_summary 中按 uuid 嵌入缩略图）
- 登录: POST {API_ROOT}/user/login（表单: username&password）→ {"code":200,"data":"<token>"}
- 全量资产列表: GET {API_ROOT}/asset?page=&pageCount=    （不带 username 即全量）
- 可选按用户过滤: GET {API_ROOT}/asset?page=&pageCount=&username=
- 资产详情: GET {API_ROOT}/asset/{uuid}
- 结构树: GET {API_ROOT}/asset/structure?uuids=<uuid>
- 导出: Excel(assets_summary / files_detail / suffix_stats / images_downloaded)
"""

import argparse
import logging
import os
import sys
import time
import urllib.parse
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ========= Session / Retry =========
def build_session(token: Optional[str] = None, retries: int = 3, backoff: float = 0.8) -> requests.Session:
    s = requests.Session()
    if token:
        s.headers["Authorization"] = f"Bearer {token}"
        s.headers["token"] = token  # 兼容部分网关
    s.headers["Accept"] = "application/json;charset=utf-8"
    s.headers["Accept-Encoding"] = "gzip"
    retry_cfg = Retry(
        total=retries, connect=retries, read=retries, status=retries,
        backoff_factor=backoff, status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset(["GET", "POST", "PUT", "DELETE", "HEAD", "OPTIONS"]),
        respect_retry_after_header=True, raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry_cfg)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    return s

# ========= Helpers =========
def _basename_from_furl(furl: Optional[str]) -> Optional[str]:
    if not furl:
        return None
    path = urllib.parse.urlparse(furl).path
    name = os.path.basename(path)
    return name or None

def _normalize_id(v: Any) -> Optional[str]:
    return str(v) if v is not None else None

def _extract_asset_list_page(body: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Optional[int], Optional[int]]:
    items: List[Dict[str, Any]] = []
    pages = body.get("pages")
    page_index = body.get("pageIndex")
    data = body.get("data")
    if isinstance(data, list):
        items = data
    elif isinstance(data, dict):
        inner = data.get("data")
        if isinstance(inner, list):
            items = inner
        pages = data.get("pages", pages)
        page_index = data.get("pageIndex", page_index)
    if not isinstance(items, list):
        items = []
    return items, pages, page_index

def _origin_from_api_root(api_root: str) -> str:
    p = urllib.parse.urlparse(api_root)
    return f"{p.scheme}://{p.netloc}"

def _to_abs_url(base_origin: str, url_or_path: Optional[str]) -> Optional[str]:
    if not url_or_path:
        return None
    u = str(url_or_path).strip()
    if u.startswith("http://") or u.startswith("https://"):
        return u
    if u.startswith("/"):
        return base_origin + u
    return base_origin + "/" + u.lstrip("/")

# ========= API calls =========
def login(api_root: str, username: str, password: str, timeout: int, sess: Optional[requests.Session] = None) -> str:
    url = api_root.rstrip("/") + "/user/login"
    s = sess or build_session()
    payload = {"username": username, "password": password}
    logging.info("POST %s (login as %s)", url, username)
    r = s.post(url, data=payload, timeout=timeout)  # 表单提交
    if not (200 <= r.status_code < 300):
        raise RuntimeError(f"HTTP {r.status_code}: {r.text}")
    body = r.json()
    if body.get("code") not in (200, 0):
        raise RuntimeError(f"登录失败：code={body.get('code')}, message={body.get('message')}")
    token = body.get("data")
    if not token or not isinstance(token, str):
        raise RuntimeError("登录成功但未返回 token（data 字段为空）")
    return token

def list_assets(api_root: str, sess: requests.Session, page_size: Optional[int] = 50,
                timeout: int = 30, owner_username: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    owner_username=None -> 拉取全量资产；否则按用户名过滤。
    """
    asset_base = api_root.rstrip("/") + "/asset"
    all_assets: List[Dict[str, Any]] = []
    page = 1
    while True:
        params = {"page": page}
        if page_size and page_size > 0:
            params["pageCount"] = page_size
        if owner_username:
            params["username"] = owner_username
        logging.info("GET %s params=%s", asset_base, params)
        r = sess.get(asset_base, params=params, timeout=timeout)
        if not (200 <= r.status_code < 300):
            raise RuntimeError(f"HTTP {r.status_code}: {r.text}")
        items, pages, page_index = _extract_asset_list_page(r.json())
        all_assets.extend(items)
        logging.debug("Fetched page %s: %d items (total so far %d)", page, len(items), len(all_assets))
        if pages is None or page_index is None or page_index >= pages:
            break
        page += 1
    return all_assets

def get_structure_filenames(api_root: str, sess: requests.Session, uuid: str,
                            timeout: int = 30) -> List[Dict[str, Any]]:
    url = api_root.rstrip("/") + "/asset/structure"
    params = [("uuids", uuid)]
    logging.info("GET %s params=%s", url, params)
    try:
        r = sess.get(url, params=params, timeout=timeout)
        if not (200 <= r.status_code < 300):
            raise RuntimeError(f"HTTP {r.status_code}: {r.text}")
        payload = r.json().get("data")
        if not payload:
            return []
        nodes: List[Dict[str, Any]] = []
        trees: List[Dict[str, Any]] = payload if isinstance(payload, list) else [payload]

        def dfs(node: Dict[str, Any]):
            for ch in node.get("children") or []:
                ch_type = ch.get("type")
                if ch_type in ("MODEL", "IMAGE"):
                    nodes.append({
                        "id": ch.get("id"),
                        "name": ch.get("name"),
                        "type": ch_type,
                        "modelType": (ch.get("details") or {}).get("modelType"),
                        "url": ch.get("url"),
                    })
                if ch.get("children"):
                    dfs(ch)

        for t in trees:
            dfs(t)
        return nodes
    except Exception as e:
        logging.warning("structure fetch failed for %s: %s", uuid, e)
        return []

def get_asset_detail(api_root: str, sess: requests.Session, uuid: str,
                     timeout: int = 30) -> Optional[Dict[str, Any]]:
    url = api_root.rstrip("/") + f"/asset/{uuid}"
    logging.info("GET %s", url)
    r = sess.get(url, timeout=timeout)
    try:
        if not (200 <= r.status_code < 300):
            raise RuntimeError(f"HTTP {r.status_code}: {r.text}")
    except Exception as e:
        logging.warning("asset detail fetch failed for %s: %s", uuid, e)
        return None
    body = r.json()
    return body.get("data", body)

# ========= Image downloader =========
_IMAGE_SUFFIXES = {"png", "jpg", "jpeg", "gif", "bmp", "webp", "svg"}

def _is_image_file_entry(file_obj: Dict[str, Any]) -> bool:
    t = (file_obj.get("type") or "").upper()
    if t == "IMAGE":
        return True
    suf = (file_obj.get("suffix") or "").lower()
    return suf in _IMAGE_SUFFIXES

def _ensure_dir(path: str):
    if path and not os.path.exists(path):
        os.makedirs(path, exist_ok=True)

def _safe_filename(name: str) -> str:
    return "".join(c if c not in r'\/:*?"<>|' else "_" for c in name)

def download_asset_images(
    api_root: str,
    sess: requests.Session,
    asset: Dict[str, Any],
    structure_files: List[Dict[str, Any]],
    images_dir: str,
) -> List[Dict[str, Any]]:
    base_origin = _origin_from_api_root(api_root)
    rows: List[Dict[str, Any]] = []
    asset_uuid = asset.get("uuid") or ""
    asset_name = asset.get("name") or ""
    _ensure_dir(images_dir)

    # 来自 asset.files
    for f in (asset.get("files") or []):
        if not _is_image_file_entry(f):
            continue
        url = _to_abs_url(base_origin, f.get("furl"))
        file_id = f.get("id")
        base = _basename_from_furl(url) or f"image_{file_id}"
        filename = _safe_filename(f"{asset_uuid}__{file_id}__{base}")
        dst = os.path.join(images_dir, filename)
        status = "ok"
        try:
            if url:
                with sess.get(url, timeout=30, stream=True) as resp:
                    if resp.status_code == 200:
                        with open(dst, "wb") as fp:
                            for chunk in resp.iter_content(chunk_size=65536):
                                if chunk:
                                    fp.write(chunk)
                    else:
                        status = f"http_{resp.status_code}"
            else:
                status = "no_url"
        except Exception as e:
            logging.warning("download image failed (files) %s: %s", url, e)
            status = f"error:{e}"
        rows.append({
            "asset_uuid": asset_uuid, "asset_name": asset_name, "source": "files",
            "file_id_or_node_id": file_id, "url": url,
            "saved_path": dst if status == "ok" else "", "status": status,
        })

    # 来自结构树 IMAGE
    for n in structure_files:
        if (n.get("type") or "").upper() != "IMAGE":
            continue
        url = _to_abs_url(base_origin, n.get("url"))
        node_id = n.get("id")
        base = _basename_from_furl(url) or n.get("name") or f"image_{node_id}"
        filename = _safe_filename(f"{asset_uuid}__{node_id}__{base}")
        dst = os.path.join(images_dir, filename)
        status = "ok"
        try:
            if url:
                with sess.get(url, timeout=30, stream=True) as resp:
                    if resp.status_code == 200:
                        with open(dst, "wb") as fp:
                            for chunk in resp.iter_content(chunk_size=65536):
                                if chunk:
                                    fp.write(chunk)
                    else:
                        status = f"http_{resp.status_code}"
            else:
                status = "no_url"
        except Exception as e:
            logging.warning("download image failed (structure) %s: %s", url, e)
            status = f"error:{e}"
        rows.append({
            "asset_uuid": asset_uuid, "asset_name": asset_name, "source": "structure",
            "file_id_or_node_id": node_id, "url": url,
            "saved_path": dst if status == "ok" else "", "status": status,
        })
    return rows

# ========= Checklist builders =========
def build_checklist_rows(asset: Dict[str, Any], structure_files: List[Dict[str, Any]],
                         detail_asset: Optional[Dict[str, Any]] = None) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    asset_name = asset.get("name")
    asset_uuid = asset.get("uuid")
    asset_class = asset.get("clasz")
    creator = asset.get("creator")
    creation_date = asset.get("creationDate")
    asset_description = (detail_asset or {}).get("description") or asset.get("description") or ""

    files = asset.get("files") or []
    rows: List[Dict[str, Any]] = []

    struct_name_by_id: Dict[str, str] = {}
    for sf in structure_files:
        sid = _normalize_id(sf.get("id"))
        if sid and sf.get("name"):
            struct_name_by_id[sid] = sf["name"]

    detail_name_by_id: Dict[str, str] = {}
    if detail_asset:
        for df in (detail_asset.get("files") or []):
            did = _normalize_id(df.get("id"))
            if not did:
                continue
            for k in ("name", "fileName", "originalName", "filename"):
                val = df.get(k)
                if val:
                    detail_name_by_id[did] = val
                    break

    def display_name(f: Dict[str, Any]) -> str:
        fid = _normalize_id(f.get("id"))
        if fid and fid in struct_name_by_id:
            return struct_name_by_id[fid]
        if fid and fid in detail_name_by_id:
            return detail_name_by_id[fid]
        base = _basename_from_furl(f.get("furl"))
        if base:
            return base
        suffix = (f.get("suffix") or "").lower() or None
        if suffix:
            return f"{f.get('type') or 'FILE'}.{suffix}"
        return f.get("type") or "FILE"

    for f in files:
        rows.append({
            "asset_name": asset_name,
            "asset_uuid": asset_uuid,
            "asset_class": asset_class,
            "creator": creator,
            "creationDate": creation_date,
            "asset_description": asset_description,
            "file_id": f.get("id"),
            "file_type": f.get("type"),
            "file_suffix": f.get("suffix"),
            "file_modelType": f.get("modelType"),
            "file_size": f.get("specification"),
            "file_url": f.get("furl"),
            "file_name": display_name(f),
        })

    type_counts: Dict[str, int] = {}
    suf_counts: Dict[str, int] = {}
    for r in rows:
        t = (r["file_type"] or "UNKNOWN").upper()
        s = (r["file_suffix"] or "NONE").lower()
        type_counts[t] = type_counts.get(t, 0) + 1
        suf_counts[s] = suf_counts.get(s, 0) + 1

    summary = {
        "asset_name": asset_name,
        "asset_uuid": asset_uuid,
        "asset_class": asset_class,
        "creator": creator,
        "creationDate": creation_date,
        "asset_description": asset_description,
        "file_count": len(rows),
        "file_types": ", ".join(f"{k}:{v}" for k, v in sorted(type_counts.items())),
        "file_suffixes": ", ".join(f"{k}:{v}" for k, v in sorted(suf_counts.items())),
    }
    return rows, summary

# ========= Main =========
def main():
    ap = argparse.ArgumentParser(description="导出系统内【所有资产】清单（可选按用户名过滤），并在 assets_summary 中嵌入缩略图。")
    ap.add_argument("--api-root", default="http://66.natic.cn:21200/v1/xasset",
                    help="API 根路径（不含 /asset /user/login）")
    # 不传 --owner = 全量；传入 --owner=username 则只拉该用户
    ap.add_argument("--owner", default=None, help="按此用户名筛选资产（缺省=全量资产）")
    ap.add_argument("--page-size", type=int, default=50)
    ap.add_argument("--timeout", type=int, default=30)
    ap.add_argument("--retries", type=int, default=3)
    ap.add_argument("--backoff", type=float, default=0.8)
    ap.add_argument("--log-level", default="INFO", help="DEBUG/INFO/WARNING/ERROR")
    ap.add_argument("--images-dir", default="asset_images", help="图片保存目录（默认 asset_images）")
    ap.add_argument("--no-download-images", action="store_true", help="不下载图片")
    ap.add_argument("--no-embed-images", action="store_true", help="不在 Excel 中嵌入缩略图")
    args = ap.parse_args()

    download_images = not args.no_download_images
    embed_images = not args.no_embed_images

    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(asctime)s | %(levelname)s | %(message)s",
    )
    api_root = args.api_root.rstrip("/")
    logging.info("API_ROOT=%s", api_root)

    login_user = input("请输入登录用户名：").strip()
    password = input("请输入密码（回显）：")

    out_in = input("输出 Excel 文件路径（默认 assets_checklist.xlsx）：").strip()
    if not out_in:
        out = "assets_checklist.xlsx"
    else:
        if out_in.endswith(("\\", "/")) or os.path.isdir(out_in):
            out = os.path.join(out_in, "assets_checklist.xlsx")
        else:
            out = out_in
    out_dir = os.path.dirname(out)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    base_sess = build_session(retries=args.retries, backoff=args.backoff)

    logging.info("[1/5] 登录为 %s", login_user)
    token = login(api_root, login_user, password, timeout=args.timeout, sess=base_sess)

    sess = build_session(token=token, retries=args.retries, backoff=args.backoff)

    scope_msg = f"全量资产" if not args.owner else f"{args.owner} 的资产"
    logging.info("[2/5] 拉取 %s（pageSize=%s）", scope_msg, args.page_size if args.page_size else "server-default")
    assets = list_assets(api_root, sess, page_size=args.page_size, timeout=args.timeout, owner_username=args.owner)
    logging.info("资产总数：%d", len(assets))

    logging.info("[3/5] 汇总每个资产的文件明细（结构树精确匹配 + 详情 fallback）")
    all_file_rows: List[Dict[str, Any]] = []
    all_summaries: List[Dict[str, Any]] = []
    images_rows: List[Dict[str, Any]] = []

    for i, a in enumerate(assets, 1):
        uuid = a.get("uuid")
        struct_files: List[Dict[str, Any]] = []
        detail_asset: Optional[Dict[str, Any]] = None
        if uuid:
            struct_files = get_structure_filenames(api_root, sess, uuid, timeout=args.timeout)
            # 结构树为空 或 列表无描述 时才补打一遍详情，减少压力
            if (not struct_files) or (not a.get("description")):
                detail_asset = get_asset_detail(api_root, sess, uuid, timeout=args.timeout)
            if download_images:
                images_rows.extend(download_asset_images(api_root, sess, a or {}, struct_files, args.images_dir))

        rows, summary = build_checklist_rows(a, struct_files, detail_asset=detail_asset)
        all_file_rows.extend(rows)
        all_summaries.append(summary)
        if i % 20 == 0:
            time.sleep(0.2)

    logging.info("[4/5] 整理数据帧与统计")
    df_files = pd.DataFrame(all_file_rows)
    df_assets = pd.DataFrame(all_summaries)
    if not df_assets.empty:
        df_assets = df_assets.sort_values(by=["file_count", "asset_name"], ascending=[False, True])
        if "thumbnail" not in df_assets.columns and embed_images:
            df_assets.insert(0, "thumbnail", "")

    logging.info("[5/5] 导出 Excel → %s", out)
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        df_assets.to_excel(xw, index=False, sheet_name="assets_summary")
        df_files.to_excel(xw, index=False, sheet_name="files_detail")

        if not df_files.empty:
            pivot_suffix = (
                df_files.assign(file_suffix=df_files["file_suffix"].fillna("NONE"))
                .pivot_table(index="file_suffix", values="file_id", aggfunc="count")
                .rename(columns={"file_id": "count"})
                .sort_values("count", ascending=False)
            )
            pivot_suffix.to_excel(xw, sheet_name="suffix_stats")

        if images_rows:
            pd.DataFrame(images_rows).to_excel(xw, index=False, sheet_name="images_downloaded")

        # 在 assets_summary 按 uuid 嵌入缩略图（默认开启）
        if embed_images and not df_assets.empty:
            try:
                from openpyxl.drawing.image import Image as XLImage
                from openpyxl.utils import get_column_letter

                ws = xw.sheets["assets_summary"]
                cols = list(df_assets.columns)
                thumb_col_idx = cols.index("thumbnail") + 1 if "thumbnail" in cols else 1
                col_letter = get_column_letter(thumb_col_idx)

                # asset_uuid -> 代表图（优先 files，其次 structure）
                rep_map: Dict[str, str] = {}
                for row in images_rows:
                    if row.get("status") == "ok" and row.get("source") == "files":
                        au = str(row.get("asset_uuid") or "")
                        sp = row.get("saved_path")
                        if au and sp and os.path.exists(sp) and au not in rep_map:
                            rep_map[au] = sp
                for row in images_rows:
                    if row.get("status") == "ok" and row.get("source") == "structure":
                        au = str(row.get("asset_uuid") or "")
                        sp = row.get("saved_path")
                        if au and sp and os.path.exists(sp) and au not in rep_map:
                            rep_map[au] = sp

                # 目标列宽
                try:
                    ws.column_dimensions[col_letter].width = 18
                except Exception:
                    pass

                # 定位 uuid 列
                try:
                    uuid_col_idx = cols.index("asset_uuid") + 1
                except ValueError:
                    uuid_col_idx = cols.index("uuid") + 1

                thumb_height = 80
                thumb_width = 80

                # 从第2行（第一行为表头）
                for excel_row in range(2, 2 + len(df_assets)):
                    au = str(ws.cell(row=excel_row, column=uuid_col_idx).value or "")
                    sp = rep_map.get(au)
                    if not sp:
                        continue
                    try:
                        img = XLImage(sp)
                        if hasattr(img, "width") and hasattr(img, "height"):
                            scale = min(thumb_width / float(img.width), thumb_height / float(img.height))
                            img.width = max(1, int(img.width * scale))
                            img.height = max(1, int(img.height * scale))
                        ws.add_image(img, f"{col_letter}{excel_row}")
                        try:
                            ws.row_dimensions[excel_row].height = max(ws.row_dimensions[excel_row].height or 0, thumb_height * 0.75)
                        except Exception:
                            pass
                    except Exception as e:
                        logging.warning("embed asset thumbnail failed for %s: %s", au, e)

            except ImportError:
                logging.warning("未安装 pillow，无法在 assets_summary 中嵌入缩略图；请执行：pip install pillow")

    logging.info("完成 ✅ 输出文件：%s", out)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logging.exception("执行失败：%s", e)
        sys.exit(1)
